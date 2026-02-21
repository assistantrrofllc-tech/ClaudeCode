"""
Dashboard routes — serves the web UI and receipt images.

All management views: home, ledger, employee management.
Receipt images served from local storage with path traversal protection.
Export endpoints: QuickBooks CSV, Google Sheets CSV, Excel (.xlsx).
"""

import csv
import io
import logging
from datetime import datetime
from pathlib import Path

from flask import (
    Blueprint, render_template, send_from_directory, jsonify, request, abort,
    Response, send_file,
)

from config.settings import RECEIPT_STORAGE_PATH
from src.database.connection import get_db

log = logging.getLogger(__name__)

dashboard_bp = Blueprint("dashboard", __name__)


# ── Pages ────────────────────────────────────────────────────


@dashboard_bp.route("/")
def home():
    """Dashboard home — spend summary, flagged receipts, recent activity."""
    db = get_db()
    try:
        stats = _get_dashboard_stats(db)
        flagged = _get_flagged_receipts(db)
        recent = _get_recent_receipts(db, limit=10)
        unknown = _get_unknown_contacts(db, limit=10)
        return render_template("index.html", stats=stats, flagged=flagged, recent=recent, unknown=unknown)
    finally:
        db.close()


# ── Receipt Image Serving ────────────────────────────────────


@dashboard_bp.route("/receipts/image/<filename>")
def serve_receipt_image(filename):
    """Serve a receipt image from local storage.

    Path traversal protection: only serves files from the receipts directory,
    filename must not contain path separators.
    """
    # Block path traversal
    if "/" in filename or "\\" in filename or ".." in filename:
        abort(404)

    storage_dir = Path(RECEIPT_STORAGE_PATH).resolve()
    file_path = (storage_dir / filename).resolve()

    # Ensure the resolved path is inside the storage directory
    if not str(file_path).startswith(str(storage_dir)):
        abort(404)

    if not file_path.exists():
        abort(404)

    return send_from_directory(str(storage_dir), filename)


# ── API Endpoints ────────────────────────────────────────────


@dashboard_bp.route("/api/receipts")
def api_receipts():
    """JSON endpoint for receipt data with filtering.

    Query params:
        period: today, week, month, ytd, all (default: all)
        start: YYYY-MM-DD custom start date
        end: YYYY-MM-DD custom end date
        employee: employee ID
        project: project ID
        vendor: vendor name (partial match)
        status: confirmed, pending, flagged
        sort: date, employee, vendor, project, amount, status (default: date)
        order: asc, desc (default: desc)
    """
    db = get_db()
    try:
        receipts = _query_receipts(db, request.args)
        return jsonify(receipts)
    finally:
        db.close()


@dashboard_bp.route("/api/receipts/export")
def api_receipts_export():
    """Export filtered receipts as QuickBooks CSV, Google Sheets CSV, or Excel.

    Uses the same filters as the main receipts API.
    Query param 'format': quickbooks, csv, excel (default: csv)
    """
    db = get_db()
    try:
        receipts = _query_receipts(db, request.args)
        fmt = request.args.get("format", "csv")

        if fmt == "quickbooks":
            return _export_quickbooks_csv(receipts)
        elif fmt == "excel":
            return _export_excel(receipts)
        else:
            return _export_csv(receipts)
    finally:
        db.close()


@dashboard_bp.route("/api/receipts/<int:receipt_id>")
def api_receipt_detail(receipt_id):
    """Single receipt with full detail including line items."""
    db = get_db()
    try:
        receipt = _get_receipt_detail(db, receipt_id)
        if not receipt:
            return jsonify({"error": "Receipt not found"}), 404
        return jsonify(receipt)
    finally:
        db.close()


@dashboard_bp.route("/api/dashboard/stats")
def api_dashboard_stats():
    """Dashboard summary stats as JSON."""
    db = get_db()
    try:
        return jsonify(_get_dashboard_stats(db))
    finally:
        db.close()


# ── Employee Management ──────────────────────────────────


@dashboard_bp.route("/employees")
def employees_page():
    """Employee management page."""
    db = get_db()
    try:
        employees = db.execute("""
            SELECT e.*,
                   (SELECT MAX(r.created_at) FROM receipts r WHERE r.employee_id = e.id) as last_submission
            FROM employees e ORDER BY e.first_name
        """).fetchall()
        return render_template("employees.html", employees=[dict(e) for e in employees])
    finally:
        db.close()


@dashboard_bp.route("/api/employees", methods=["GET"])
def api_employees():
    """List all employees as JSON."""
    db = get_db()
    try:
        rows = db.execute("""
            SELECT e.*,
                   (SELECT MAX(r.created_at) FROM receipts r WHERE r.employee_id = e.id) as last_submission
            FROM employees e ORDER BY e.first_name
        """).fetchall()
        return jsonify([dict(r) for r in rows])
    finally:
        db.close()


@dashboard_bp.route("/api/employees", methods=["POST"])
def api_add_employee():
    """Add a new employee."""
    data = request.get_json()
    if not data or not data.get("first_name") or not data.get("phone_number"):
        return jsonify({"error": "first_name and phone_number are required"}), 400

    phone = data["phone_number"].strip()
    if not phone.startswith("+"):
        phone = "+1" + phone.replace("-", "").replace(" ", "").replace("(", "").replace(")", "")

    db = get_db()
    try:
        existing = db.execute("SELECT id FROM employees WHERE phone_number = ?", (phone,)).fetchone()
        if existing:
            return jsonify({"error": "Phone number already registered"}), 409

        db.execute(
            "INSERT INTO employees (phone_number, first_name, full_name, role, crew) VALUES (?, ?, ?, ?, ?)",
            (phone, data["first_name"], data.get("full_name"), data.get("role"), data.get("crew")),
        )
        db.commit()
        return jsonify({"status": "created", "phone_number": phone}), 201
    finally:
        db.close()


@dashboard_bp.route("/api/employees/<int:employee_id>", methods=["GET"])
def api_employee_detail(employee_id):
    """Get a single employee (also serves as CrewCert QR landing page)."""
    db = get_db()
    try:
        row = db.execute("SELECT * FROM employees WHERE id = ?", (employee_id,)).fetchone()
        if not row:
            return jsonify({"error": "Employee not found"}), 404
        return jsonify(dict(row))
    finally:
        db.close()


@dashboard_bp.route("/api/employees/<int:employee_id>", methods=["PUT"])
def api_update_employee(employee_id):
    """Update employee fields."""
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    allowed = {"first_name", "full_name", "role", "crew", "phone_number"}
    updates = {k: v for k, v in data.items() if k in allowed}
    if not updates:
        return jsonify({"error": "No valid fields to update"}), 400

    set_clause = ", ".join(f"{k} = ?" for k in updates)
    values = list(updates.values()) + [employee_id]

    db = get_db()
    try:
        db.execute(f"UPDATE employees SET {set_clause}, updated_at = datetime('now') WHERE id = ?", values)
        db.commit()
        return jsonify({"status": "updated"})
    finally:
        db.close()


@dashboard_bp.route("/api/employees/<int:employee_id>/deactivate", methods=["POST"])
def api_deactivate_employee(employee_id):
    """Deactivate an employee — they can no longer submit receipts."""
    db = get_db()
    try:
        db.execute("UPDATE employees SET is_active = 0, updated_at = datetime('now') WHERE id = ?", (employee_id,))
        db.commit()
        return jsonify({"status": "deactivated"})
    finally:
        db.close()


@dashboard_bp.route("/api/employees/<int:employee_id>/activate", methods=["POST"])
def api_activate_employee(employee_id):
    """Reactivate an employee."""
    db = get_db()
    try:
        db.execute("UPDATE employees SET is_active = 1, updated_at = datetime('now') WHERE id = ?", (employee_id,))
        db.commit()
        return jsonify({"status": "activated"})
    finally:
        db.close()


# ── Unknown Contacts ─────────────────────────────────────


@dashboard_bp.route("/api/unknown-contacts")
def api_unknown_contacts():
    """List recent unknown contact attempts."""
    db = get_db()
    try:
        rows = db.execute("""
            SELECT * FROM unknown_contacts ORDER BY created_at DESC LIMIT 50
        """).fetchall()
        return jsonify([dict(r) for r in rows])
    finally:
        db.close()


# ── Ledger Page ──────────────────────────────────────────


@dashboard_bp.route("/ledger")
def ledger_page():
    """Banking-style transaction ledger — Kim's primary view."""
    db = get_db()
    try:
        employees = db.execute("SELECT id, first_name FROM employees ORDER BY first_name").fetchall()
        projects = db.execute("SELECT id, name FROM projects WHERE status = 'active' ORDER BY name").fetchall()
        return render_template(
            "ledger.html",
            employees=[dict(e) for e in employees],
            projects=[dict(p) for p in projects],
        )
    finally:
        db.close()


# ── Email Settings ──────────────────────────────────────


@dashboard_bp.route("/settings")
def settings_page():
    """Email settings page."""
    db = get_db()
    try:
        rows = db.execute("SELECT key, value FROM email_settings").fetchall()
        settings = {r["key"]: r["value"] for r in rows}
        employees = db.execute("SELECT id, first_name FROM employees ORDER BY first_name").fetchall()
        projects = db.execute("SELECT id, name FROM projects WHERE status = 'active' ORDER BY name").fetchall()
        return render_template(
            "settings.html",
            settings=settings,
            employees=[dict(e) for e in employees],
            projects=[dict(p) for p in projects],
        )
    finally:
        db.close()


@dashboard_bp.route("/api/settings", methods=["GET"])
def api_get_settings():
    """Get all email settings."""
    db = get_db()
    try:
        rows = db.execute("SELECT key, value FROM email_settings").fetchall()
        return jsonify({r["key"]: r["value"] for r in rows})
    finally:
        db.close()


@dashboard_bp.route("/api/settings", methods=["PUT"])
def api_update_settings():
    """Update email settings."""
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    allowed_keys = {
        "recipient_email", "frequency", "day_of_week", "time_of_day",
        "include_scope", "include_filter", "enabled",
    }

    db = get_db()
    try:
        for key, value in data.items():
            if key in allowed_keys:
                db.execute(
                    "INSERT OR REPLACE INTO email_settings (key, value, updated_at) VALUES (?, ?, datetime('now'))",
                    (key, str(value)),
                )
        db.commit()
        return jsonify({"status": "updated"})
    finally:
        db.close()


@dashboard_bp.route("/api/settings/send-now", methods=["POST"])
def api_send_report_now():
    """Trigger an immediate email report with current settings."""
    db = get_db()
    try:
        rows = db.execute("SELECT key, value FROM email_settings").fetchall()
        settings = {r["key"]: r["value"] for r in rows}
        recipient = settings.get("recipient_email", "")
        if not recipient:
            return jsonify({"error": "No recipient email configured"}), 400

        # Trigger the existing weekly report send endpoint
        from flask import current_app
        with current_app.test_client() as client:
            resp = client.post(f"/reports/weekly/send?recipient={recipient}")
            if resp.status_code == 200:
                return jsonify({"status": "sent", "recipient": recipient})
            return jsonify({"error": "Failed to send report"}), 500
    finally:
        db.close()


# ── Export Helpers ────────────────────────────────────────────


def _export_csv(receipts: list) -> Response:
    """Export as standard CSV (Google Sheets compatible)."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Employee", "Vendor", "Project", "Subtotal", "Tax", "Total", "Payment Method", "Status"])
    for r in receipts:
        writer.writerow([
            r.get("purchase_date", ""),
            r.get("employee_name", ""),
            r.get("vendor_name", ""),
            r.get("project_name") or r.get("matched_project_name", ""),
            r.get("subtotal", ""),
            r.get("tax", ""),
            r.get("total", ""),
            r.get("payment_method", ""),
            r.get("status", ""),
        ])

    resp = Response(output.getvalue(), mimetype="text/csv")
    resp.headers["Content-Disposition"] = f"attachment; filename=crewledger_export_{datetime.now().strftime('%Y%m%d')}.csv"
    return resp


def _export_quickbooks_csv(receipts: list) -> Response:
    """Export as QuickBooks IIF/CSV format for expense import."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Vendor", "Account", "Amount", "Memo", "Payment Method"])
    for r in receipts:
        project = r.get("project_name") or r.get("matched_project_name", "")
        memo = f"Employee: {r.get('employee_name', '')} | Project: {project}"
        writer.writerow([
            r.get("purchase_date", ""),
            r.get("vendor_name", ""),
            "Materials & Supplies",
            r.get("total", ""),
            memo,
            r.get("payment_method", ""),
        ])

    resp = Response(output.getvalue(), mimetype="text/csv")
    resp.headers["Content-Disposition"] = f"attachment; filename=crewledger_quickbooks_{datetime.now().strftime('%Y%m%d')}.csv"
    return resp


def _export_excel(receipts: list) -> Response:
    """Export as Excel (.xlsx) with formatting."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CrewLedger Export"

    # Header row
    headers = ["Date", "Employee", "Vendor", "Project", "Subtotal", "Tax", "Total", "Payment Method", "Status"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, r in enumerate(receipts, 2):
        ws.cell(row=row_idx, column=1, value=r.get("purchase_date", ""))
        ws.cell(row=row_idx, column=2, value=r.get("employee_name", ""))
        ws.cell(row=row_idx, column=3, value=r.get("vendor_name", ""))
        ws.cell(row=row_idx, column=4, value=r.get("project_name") or r.get("matched_project_name", ""))
        ws.cell(row=row_idx, column=5, value=r.get("subtotal") or 0).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=6, value=r.get("tax") or 0).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=7, value=r.get("total") or 0).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=8, value=r.get("payment_method", ""))
        ws.cell(row=row_idx, column=9, value=r.get("status", ""))

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    # Total row
    total_row = len(receipts) + 2
    ws.cell(row=total_row, column=6, value="TOTAL:").font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=sum(r.get("total", 0) or 0 for r in receipts)).font = Font(bold=True)
    ws.cell(row=total_row, column=7).number_format = '#,##0.00'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"crewledger_export_{datetime.now().strftime('%Y%m%d')}.xlsx",
    )


# ── Data Helpers ─────────────────────────────────────────────


def _get_dashboard_stats(db) -> dict:
    """Summary stats for the dashboard home screen."""
    row = db.execute("""
        SELECT
            COALESCE(SUM(CASE WHEN created_at >= date('now', 'weekday 1', '-7 days') THEN total ELSE 0 END), 0) as week_spend,
            COALESCE(SUM(CASE WHEN created_at >= date('now', 'start of month') THEN total ELSE 0 END), 0) as month_spend,
            COUNT(*) as total_receipts,
            SUM(CASE WHEN status = 'flagged' THEN 1 ELSE 0 END) as flagged_count,
            SUM(CASE WHEN status = 'pending' THEN 1 ELSE 0 END) as pending_count,
            SUM(CASE WHEN status = 'confirmed' THEN 1 ELSE 0 END) as confirmed_count
        FROM receipts
    """).fetchone()

    employee_count = db.execute("SELECT COUNT(*) as cnt FROM employees WHERE is_active = 1").fetchone()["cnt"]
    project_count = db.execute("SELECT COUNT(*) as cnt FROM projects WHERE status = 'active'").fetchone()["cnt"]

    unknown_count = db.execute("SELECT COUNT(*) as cnt FROM unknown_contacts").fetchone()["cnt"]

    return {
        "week_spend": round(row["week_spend"], 2),
        "month_spend": round(row["month_spend"], 2),
        "total_receipts": row["total_receipts"],
        "flagged_count": row["flagged_count"],
        "pending_count": row["pending_count"],
        "confirmed_count": row["confirmed_count"],
        "employee_count": employee_count,
        "project_count": project_count,
        "unknown_count": unknown_count,
    }


def _get_flagged_receipts(db, limit=20) -> list:
    """Flagged receipts for the review queue."""
    rows = db.execute("""
        SELECT r.*, e.first_name as employee_name, p.name as project_name
        FROM receipts r
        LEFT JOIN employees e ON r.employee_id = e.id
        LEFT JOIN projects p ON r.project_id = p.id
        WHERE r.status = 'flagged'
        ORDER BY r.created_at DESC
        LIMIT ?
    """, (limit,)).fetchall()
    return [_row_to_dict(r) for r in rows]


def _get_recent_receipts(db, limit=10) -> list:
    """Most recent receipts for the activity feed."""
    rows = db.execute("""
        SELECT r.*, e.first_name as employee_name, p.name as project_name
        FROM receipts r
        LEFT JOIN employees e ON r.employee_id = e.id
        LEFT JOIN projects p ON r.project_id = p.id
        ORDER BY r.created_at DESC
        LIMIT ?
    """, (limit,)).fetchall()
    return [_row_to_dict(r) for r in rows]


def _query_receipts(db, args) -> list:
    """Query receipts with filters and sorting."""
    conditions = []
    params = []

    # Period filter
    period = args.get("period", "all")
    if period == "today":
        conditions.append("r.created_at >= date('now')")
    elif period == "week":
        conditions.append("r.created_at >= date('now', 'weekday 1', '-7 days')")
    elif period == "month":
        conditions.append("r.created_at >= date('now', 'start of month')")
    elif period == "ytd":
        conditions.append("r.created_at >= date('now', 'start of year')")

    # Custom date range
    start = args.get("start")
    end = args.get("end")
    if start:
        conditions.append("r.created_at >= ?")
        params.append(start)
    if end:
        conditions.append("r.created_at < date(?, '+1 day')")
        params.append(end)

    # Filters
    employee_id = args.get("employee")
    if employee_id:
        conditions.append("r.employee_id = ?")
        params.append(employee_id)

    project_id = args.get("project")
    if project_id:
        conditions.append("r.project_id = ?")
        params.append(project_id)

    vendor = args.get("vendor")
    if vendor:
        conditions.append("r.vendor_name LIKE ?")
        params.append(f"%{vendor}%")

    status = args.get("status")
    if status:
        conditions.append("r.status = ?")
        params.append(status)

    where = " AND ".join(conditions) if conditions else "1=1"

    # Sorting
    sort_map = {
        "date": "r.created_at",
        "employee": "e.first_name",
        "vendor": "r.vendor_name",
        "project": "p.name",
        "amount": "r.total",
        "status": "r.status",
        "category": "r.vendor_name",
    }
    sort_col = sort_map.get(args.get("sort", "date"), "r.created_at")
    order = "ASC" if args.get("order") == "asc" else "DESC"

    rows = db.execute(f"""
        SELECT r.*, e.first_name as employee_name, e.crew,
               p.name as project_name
        FROM receipts r
        LEFT JOIN employees e ON r.employee_id = e.id
        LEFT JOIN projects p ON r.project_id = p.id
        WHERE {where}
        ORDER BY {sort_col} {order}
        LIMIT 500
    """, params).fetchall()

    return [_row_to_dict(r) for r in rows]


def _get_receipt_detail(db, receipt_id: int) -> dict | None:
    """Single receipt with line items."""
    row = db.execute("""
        SELECT r.*, e.first_name as employee_name, e.crew,
               p.name as project_name
        FROM receipts r
        LEFT JOIN employees e ON r.employee_id = e.id
        LEFT JOIN projects p ON r.project_id = p.id
        WHERE r.id = ?
    """, (receipt_id,)).fetchone()

    if not row:
        return None

    result = _row_to_dict(row)

    items = db.execute("""
        SELECT li.*, c.name as category_name
        FROM line_items li
        LEFT JOIN categories c ON li.category_id = c.id
        WHERE li.receipt_id = ?
        ORDER BY li.id
    """, (receipt_id,)).fetchall()

    result["line_items"] = [dict(i) for i in items]
    return result


def _get_unknown_contacts(db, limit=10) -> list:
    """Recent unknown contact attempts for dashboard."""
    rows = db.execute("""
        SELECT * FROM unknown_contacts ORDER BY created_at DESC LIMIT ?
    """, (limit,)).fetchall()
    return [dict(r) for r in rows]


def _row_to_dict(row) -> dict:
    """Convert a sqlite3.Row to a plain dict."""
    d = dict(row)
    # Add image URL if image exists
    if d.get("image_path"):
        filename = Path(d["image_path"]).name
        d["image_url"] = f"/receipts/image/{filename}"
    else:
        d["image_url"] = None
    return d

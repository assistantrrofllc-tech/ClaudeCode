#!/usr/bin/env python3
"""
Import vehicles and maintenance records from the RROF fleet spreadsheet.

Reads the Excel workbook at the given path (default: ~/Desktop/RROF - TRUCK - VANS - TRAILERS.xlsx),
creates the vehicles and vehicle_maintenance tables if they don't exist, and inserts
vehicle info + maintenance history from each sheet.

Usage:
    python scripts/import_vehicles.py
    python scripts/import_vehicles.py --file /path/to/spreadsheet.xlsx
    python scripts/import_vehicles.py --db data/crewledger.db
    python scripts/import_vehicles.py --dry-run

Safe to run multiple times -- vehicles are matched by VIN (or year+make+model
if VIN is missing) and skipped if already present.
"""

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path

# Allow imports from project root
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.database.connection import get_db

DEFAULT_XLSX = os.path.expanduser(
    "~/Desktop/RROF - TRUCK - VANS - TRAILERS.xlsx"
)

# Sheets to skip entirely (not vehicle sheets)
SKIP_SHEETS = {"INSURANCE"}

# ------------------------------------------------------------------ #
#  Vehicle metadata overrides keyed by sheet name.                    #
#  The spreadsheet's own G/H cells are the primary source; these      #
#  overrides fill in color and assigned_to parsed from the sheet name, #
#  plus corrections where the sheet title disagrees with the cell data #
# ------------------------------------------------------------------ #
SHEET_META = {
    "2009 F150 -  ":                      {"color": None,    "assigned_to": None},
    "2015 Van E350 - JUSTINO":            {"color": None,    "assigned_to": "Justino"},
    "2019 Transit Van - BLACK - OMAR":    {"color": "Black", "assigned_to": "Omar"},
    "2009 F250 - SANTOS":                 {"color": None,    "assigned_to": "Santos"},
    "2021 Tahoe - white":                 {"color": "White", "assigned_to": None},
    "2023 Tahoe-green":                   {"color": "Green", "assigned_to": None},
    "2016 Transit Van - Mario":           {"color": None,    "assigned_to": "Mario"},
    "2017 Transit Van - Wilson":          {"color": None,    "assigned_to": "Wilson"},
    "2020 F150 - Blue ":                  {"color": "Blue",  "assigned_to": None},
    "2025 Jaguar":                        {"color": None,    "assigned_to": None},
    "2023 F150 - Grey":                   {"color": "Grey",  "assigned_to": None},
    "2003 F250 - Silver ":                {"color": "Silver","assigned_to": None},
    "2020 F150 - Red":                    {"color": "Red",   "assigned_to": None},
    "2018 F150 - BLACK -":                {"color": "Black", "assigned_to": "Robert"},
    "GRAY TRAILER":                       {"color": "Gray",  "assigned_to": None},
    "BLACK TRAILER":                      {"color": "Black", "assigned_to": None},
    "2011 E350 - EXTRA VAN -SOLD":        {"color": None,    "assigned_to": None},
    "Tahoe - blue -SOLD":                 {"color": "Blue",  "assigned_to": None},
    "2011 E350 - Mario - SOLD":           {"color": None,    "assigned_to": "Mario"},
}


# ------------------------------------------------------------------ #
#  Table creation                                                     #
# ------------------------------------------------------------------ #

VEHICLES_DDL = """
CREATE TABLE IF NOT EXISTS vehicles (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    year            INTEGER,
    make            TEXT,
    model           TEXT,
    color           TEXT,
    tire_size       TEXT,
    plate_number    TEXT,
    vin             TEXT    UNIQUE,
    nickname        TEXT,
    assigned_to     TEXT,
    status          TEXT    DEFAULT 'active'
                           CHECK(status IN ('active', 'sold', 'out_of_service')),
    created_at      TEXT    DEFAULT (datetime('now')),
    updated_at      TEXT    DEFAULT (datetime('now'))
);
"""

VEHICLES_INDEXES = [
    "CREATE INDEX IF NOT EXISTS idx_vehicles_status ON vehicles(status);",
    "CREATE INDEX IF NOT EXISTS idx_vehicles_vin    ON vehicles(vin);",
]

MAINTENANCE_DDL = """
CREATE TABLE IF NOT EXISTS vehicle_maintenance (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    vehicle_id      INTEGER NOT NULL,
    service_date    TEXT,
    description     TEXT,
    cost            REAL,
    mileage         INTEGER,
    vendor          TEXT,
    created_at      TEXT    DEFAULT (datetime('now')),
    FOREIGN KEY (vehicle_id) REFERENCES vehicles(id) ON DELETE CASCADE
);
"""

MAINTENANCE_INDEXES = [
    "CREATE INDEX IF NOT EXISTS idx_veh_maint_vehicle ON vehicle_maintenance(vehicle_id);",
    "CREATE INDEX IF NOT EXISTS idx_veh_maint_date    ON vehicle_maintenance(service_date);",
]


def ensure_tables(db):
    """Create vehicles and vehicle_maintenance tables if they don't exist."""
    db.executescript(VEHICLES_DDL)
    for idx in VEHICLES_INDEXES:
        db.execute(idx)
    db.executescript(MAINTENANCE_DDL)
    for idx in MAINTENANCE_INDEXES:
        db.execute(idx)
    db.commit()


# ------------------------------------------------------------------ #
#  Helpers                                                            #
# ------------------------------------------------------------------ #

def clean_str(val):
    """Return a stripped string or None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def parse_year(val):
    """Extract an integer year from a cell value."""
    if val is None:
        return None
    try:
        y = int(float(val))
        return y if 1900 <= y <= 2100 else None
    except (ValueError, TypeError):
        return None


def parse_cost(val):
    """Parse a cost value that might be a float, int, or messy string."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return round(float(val), 2)
    s = str(val).strip()
    if not s:
        return None
    # Strip dollar signs, commas, spaces
    s = s.replace("$", "").replace(",", "").replace(" ", "")
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def parse_mileage(val):
    """Parse mileage that might be a float, int, or messy string."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        try:
            m = int(float(val))
            return m if m > 0 else None
        except (ValueError, TypeError):
            return None
    s = str(val).strip().replace(",", "").replace(" ", "")
    if not s:
        return None
    try:
        m = int(float(s))
        return m if m > 0 else None
    except ValueError:
        return None


def parse_date(val):
    """Parse a date from a datetime object or string. Returns ISO date string or None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    # Try common formats
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def is_sold(sheet_name):
    """Check if a sheet name indicates the vehicle has been sold."""
    return "SOLD" in sheet_name.upper()


# ------------------------------------------------------------------ #
#  Sheet parsing                                                      #
# ------------------------------------------------------------------ #

def parse_vehicle_info(ws, sheet_name):
    """
    Extract vehicle metadata from columns G-H (rows 2-8).

    Layout:
        Row 1: G="Vehicle Info"
        Row 2: G="Year",       H=<year>
        Row 3: G="Make",       H=<make>
        Row 4: G="Model",      H=<model>
        Row 5: G="Tire size",  H=<tire_size>
        Row 6: G="Engine size",H=<engine_size>  (unused)
        Row 7: G="FL Plate#",  H=<plate>
        Row 8: G="VIN#",       H=<vin>
    """
    year = parse_year(ws.cell(row=2, column=8).value)
    make = clean_str(ws.cell(row=3, column=8).value)
    model = clean_str(ws.cell(row=4, column=8).value)
    tire_size = clean_str(ws.cell(row=5, column=8).value)
    plate = clean_str(ws.cell(row=7, column=8).value)
    vin = clean_str(ws.cell(row=8, column=8).value)

    # Tire size of "0" or "0.0" means unknown
    if tire_size and tire_size in ("0", "0.0"):
        tire_size = None

    meta = SHEET_META.get(sheet_name, {})
    color = meta.get("color")
    assigned_to = meta.get("assigned_to")
    status = "sold" if is_sold(sheet_name) else "active"

    return {
        "year": year,
        "make": make.title() if make else None,
        "model": model,
        "color": color,
        "tire_size": tire_size,
        "plate_number": plate,
        "vin": vin,
        "nickname": sheet_name.strip(),
        "assigned_to": assigned_to,
        "status": status,
    }


def parse_maintenance_records(ws):
    """
    Extract maintenance records from columns A-E starting at row 2.

    Layout:
        Row 1: headers (Date, Repairs, Cost, Mileage, [Vendor])
        Row 2+: data

    Some rows have a description continuation (date is None but description
    is present). Those get appended to the previous record's description.
    """
    records = []
    pending = None  # track multi-line descriptions

    for row_idx in range(2, 500):
        date_val = ws.cell(row=row_idx, column=1).value
        desc_val = clean_str(ws.cell(row=row_idx, column=2).value)
        cost_val = ws.cell(row=row_idx, column=3).value
        mile_val = ws.cell(row=row_idx, column=4).value
        vendor_val = clean_str(ws.cell(row=row_idx, column=5).value)

        date_parsed = parse_date(date_val)
        cost_parsed = parse_cost(cost_val)
        mile_parsed = parse_mileage(mile_val)

        has_date = date_parsed is not None
        has_desc = desc_val is not None
        has_cost = cost_parsed is not None

        # Completely empty row -- check a few more before giving up
        if not has_date and not has_desc and not has_cost:
            # Flush pending
            if pending:
                records.append(pending)
                pending = None
            # Look ahead for more data
            more_data = False
            for lookahead in range(row_idx + 1, min(row_idx + 5, 500)):
                if ws.cell(row=lookahead, column=1).value or \
                   clean_str(ws.cell(row=lookahead, column=2).value):
                    more_data = True
                    break
            if not more_data:
                break
            continue

        # Continuation row: no date, but has description
        # (and no cost of its own) -- append to previous
        if not has_date and has_desc and not has_cost and pending:
            pending["description"] += "; " + desc_val
            continue

        # New record -- flush previous first
        if pending:
            records.append(pending)

        # Skip vendor column values that look like trade-in notes rather than vendors
        if vendor_val and vendor_val.startswith("-"):
            vendor_val = None

        pending = {
            "service_date": date_parsed,
            "description": desc_val,
            "cost": cost_parsed,
            "mileage": mile_parsed,
            "vendor": vendor_val,
        }

    # Flush last pending record
    if pending:
        records.append(pending)

    return records


# ------------------------------------------------------------------ #
#  Import logic                                                       #
# ------------------------------------------------------------------ #

def find_existing_vehicle(db, vin, year, make, model):
    """
    Check if a vehicle already exists. Match by VIN first (most reliable),
    fall back to year+make+model.
    """
    if vin:
        row = db.execute(
            "SELECT id FROM vehicles WHERE vin = ?", (vin,)
        ).fetchone()
        if row:
            return row["id"]
    # Fallback: year + make + model (case-insensitive)
    if year and make and model:
        row = db.execute(
            """SELECT id FROM vehicles
               WHERE year = ? AND LOWER(make) = LOWER(?) AND LOWER(model) = LOWER(?)""",
            (year, make, model),
        ).fetchone()
        if row:
            return row["id"]
    return None


def import_vehicles(db_path, xlsx_path, dry_run=False):
    """Main import routine."""
    try:
        import openpyxl
    except ImportError:
        print("Error: openpyxl is required. Install with: pip install openpyxl")
        sys.exit(1)

    if not os.path.exists(xlsx_path):
        print(f"Error: File not found: {xlsx_path}")
        sys.exit(1)

    db = get_db(db_path)
    try:
        ensure_tables(db)

        wb = openpyxl.load_workbook(xlsx_path, data_only=True)

        vehicles_inserted = 0
        vehicles_skipped = 0
        maintenance_inserted = 0
        maintenance_skipped = 0
        errors = []

        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_SHEETS:
                print(f"  SKIP  {sheet_name} (reference sheet)")
                continue

            ws = wb[sheet_name]

            # --- Vehicle info ---
            try:
                info = parse_vehicle_info(ws, sheet_name)
            except Exception as e:
                errors.append(f"  {sheet_name}: failed to parse vehicle info: {e}")
                continue

            if not info["make"] and not info["model"]:
                errors.append(f"  {sheet_name}: no make/model found, skipping")
                continue

            existing_id = find_existing_vehicle(
                db, info["vin"], info["year"], info["make"], info["model"]
            )

            if existing_id:
                vehicle_id = existing_id
                vehicles_skipped += 1
                label = "EXISTS"
            else:
                if dry_run:
                    vehicle_id = None
                    vehicles_inserted += 1
                    label = "DRY-RUN"
                else:
                    cur = db.execute(
                        """INSERT INTO vehicles
                           (year, make, model, color, tire_size, plate_number,
                            vin, nickname, assigned_to, status)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (
                            info["year"], info["make"], info["model"],
                            info["color"], info["tire_size"], info["plate_number"],
                            info["vin"], info["nickname"], info["assigned_to"],
                            info["status"],
                        ),
                    )
                    vehicle_id = cur.lastrowid
                    vehicles_inserted += 1
                    label = "INSERT"

            # --- Maintenance records ---
            try:
                records = parse_maintenance_records(ws)
            except Exception as e:
                errors.append(f"  {sheet_name}: failed to parse maintenance: {e}")
                records = []

            maint_count = 0
            for rec in records:
                # Skip completely empty records (no description and no cost)
                if not rec["description"] and rec["cost"] is None:
                    continue

                if vehicle_id and not dry_run:
                    # Check for duplicate: same vehicle + date + description
                    dup = db.execute(
                        """SELECT id FROM vehicle_maintenance
                           WHERE vehicle_id = ?
                             AND service_date IS ?
                             AND description IS ?""",
                        (vehicle_id, rec["service_date"], rec["description"]),
                    ).fetchone()
                    if dup:
                        maintenance_skipped += 1
                        continue

                    db.execute(
                        """INSERT INTO vehicle_maintenance
                           (vehicle_id, service_date, description, cost, mileage, vendor)
                           VALUES (?, ?, ?, ?, ?, ?)""",
                        (
                            vehicle_id,
                            rec["service_date"],
                            rec["description"],
                            rec["cost"],
                            rec["mileage"],
                            rec["vendor"],
                        ),
                    )
                maintenance_inserted += 1
                maint_count += 1

            status_tag = f"[{info['status'].upper()}]" if info["status"] == "sold" else ""
            assigned_tag = f" -> {info['assigned_to']}" if info["assigned_to"] else ""
            print(
                f"  {label:>8}  {info['year']} {info['make']} {info['model']}"
                f"{assigned_tag} {status_tag} ({maint_count} maint records)"
            )

        if not dry_run:
            db.commit()

        # --- Summary ---
        print()
        print("=" * 50)
        print("Vehicle Import Summary")
        print("=" * 50)
        print(f"  Vehicles inserted:      {vehicles_inserted}")
        print(f"  Vehicles skipped (dup): {vehicles_skipped}")
        print(f"  Maintenance inserted:   {maintenance_inserted}")
        print(f"  Maintenance skipped:    {maintenance_skipped}")
        if dry_run:
            print("  ** DRY RUN -- no data was written **")
        if errors:
            print()
            print("Errors:")
            for e in errors:
                print(f"  {e}")
        print()

    finally:
        db.close()


# ------------------------------------------------------------------ #
#  CLI entry point                                                    #
# ------------------------------------------------------------------ #

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Import RROF fleet vehicles and maintenance from Excel"
    )
    parser.add_argument(
        "--file",
        default=DEFAULT_XLSX,
        help=f"Path to Excel workbook (default: {DEFAULT_XLSX})",
    )
    parser.add_argument(
        "--db",
        default=None,
        help="Path to database (default: from env or data/crewledger.db)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and print without writing to the database",
    )
    args = parser.parse_args()

    print(f"Importing vehicles from: {args.file}")
    print(f"Database: {args.db or os.getenv('DATABASE_PATH', 'data/crewledger.db')}")
    if args.dry_run:
        print("** DRY RUN MODE **")
    print()

    import_vehicles(args.db, args.file, dry_run=args.dry_run)

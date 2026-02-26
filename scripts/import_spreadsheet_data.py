"""
Import data from MISC ITEMS.xlsx and 2026 SCHEDULE.xlsx into CrewLedger.

Imports:
1. PERDIEM-SHIRTS-EMAILS  -> Update employees with email/phone
2. Material Suppliers      -> vendors table (created if needed)
3. PROJECT ASSIGNMENTS     -> project_assignments table (created if needed)
4. SAFETY TRAINING         -> certifications table (update existing)
5. SHOP INVENTORY (Orlando)-> inventory table (created if needed)
6. TAMPA SHOP INVENTORY    -> inventory table (created if needed)

Safe to run multiple times — checks before insert, only fills blanks on updates.

Run with: python scripts/import_spreadsheet_data.py
"""

import os
import re
import sys
from pathlib import Path

# Allow imports from project root
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from thefuzz import fuzz

from src.database.connection import get_db
from src.messaging.sms_handler import normalize_phone

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
MISC_ITEMS_PATH = os.path.expanduser("~/Desktop/MISC ITEMS.xlsx")
SCHEDULE_PATH = os.path.expanduser("~/Desktop/2026 SCHEDULE.xlsx")

# Fuzzy match threshold — 70 is lenient enough for nickname/partial matches
FUZZY_THRESHOLD = 70

# Map spreadsheet safety training columns to certification_types slugs
SAFETY_CERT_MAP = {
    "OSHA 10": "osha-10",
    "OSHA 30": "osha-30",
    "FALL PROTECTION": "fall-protection",
    "LULL/FORK LIFT": "ext-reach-forklift",
    "MANLIFT": "aerial-work-platform",
    "FIRST AID": "first-aid-cpr",
    "SIGNALMAN": "rigger-signal-person",
    "RIGGER": "basic-rigging",
}

# Safety training column letters in the spreadsheet (row 2 header)
SAFETY_COLUMNS = {
    "C": "OSHA 10",
    "D": "OSHA 30",
    "E": "FALL PROTECTION",
    "F": "LULL/FORK LIFT",
    "G": "MANLIFT",
    "H": "FIRST AID",
    "I": "SIGNALMAN",
    "J": "RIGGER",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def clean(val):
    """Return a cleaned string or None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def clean_name(val):
    """Clean a name: strip whitespace, remove trailing annotations like (D), (Cat)."""
    s = clean(val)
    if not s:
        return None
    # Remove trailing parenthetical notes like (D), (Cat), (Cuba)
    s = re.sub(r"\s*\([^)]*\)\s*$", "", s).strip()
    return s if s else None


def normalize_phone_value(val):
    """Handle phone numbers that may be stored as floats, strings, etc."""
    if val is None:
        return None
    # openpyxl sometimes reads phone numbers as floats (e.g. 5014789840.0)
    if isinstance(val, float):
        val = str(int(val))
    val = str(val).strip()
    if not val:
        return None
    return normalize_phone(val)


def load_employees(db):
    """Load all active employees from DB as list of dicts."""
    # Check which columns exist (VPS has email, local may not)
    cols_info = db.execute("PRAGMA table_info(employees)").fetchall()
    col_names = {c[1] for c in cols_info}
    select_cols = ["id", "first_name", "full_name", "phone_number"]
    if "email" in col_names:
        select_cols.append("email")
    rows = db.execute(
        f"SELECT {', '.join(select_cols)} FROM employees WHERE is_active = 1"
    ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        if "email" not in d:
            d["email"] = None
        result.append(d)
    return result


def fuzzy_match_employee(name, employees):
    """Find the best matching employee by name. Returns (employee_dict, score) or (None, 0)."""
    if not name:
        return None, 0

    name_clean = clean_name(name)
    if not name_clean:
        return None, 0

    best_match = None
    best_score = 0

    for emp in employees:
        # Try matching against full_name first, then first_name
        candidates = []
        if emp["full_name"]:
            candidates.append(emp["full_name"])
        if emp["first_name"]:
            candidates.append(emp["first_name"])

        for candidate in candidates:
            # Use token_sort_ratio for better matching with name order differences
            score = fuzz.token_sort_ratio(name_clean.lower(), candidate.lower())
            if score > best_score:
                best_score = score
                best_match = emp

            # Also try partial_ratio for substring matches (e.g. "Mario" vs "Mario Martinez")
            partial = fuzz.partial_ratio(name_clean.lower(), candidate.lower())
            if partial > best_score:
                best_score = partial
                best_match = emp

    if best_score >= FUZZY_THRESHOLD:
        return best_match, best_score
    return None, best_score


def safe_cell_dict(row):
    """Build {column_letter: value} dict, skipping MergedCell objects."""
    result = {}
    for c in row:
        try:
            result[c.column_letter] = c.value
        except AttributeError:
            pass  # MergedCell — no column_letter
    return result


def ensure_table(db, ddl):
    """Execute CREATE TABLE IF NOT EXISTS statement."""
    db.execute(ddl)
    db.commit()


# ---------------------------------------------------------------------------
# Table creation DDL
# ---------------------------------------------------------------------------
VENDORS_DDL = """
CREATE TABLE IF NOT EXISTS vendors (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    name            TEXT    UNIQUE NOT NULL,
    address         TEXT,
    phone           TEXT,
    contact_name    TEXT,
    items           TEXT,
    has_account     INTEGER DEFAULT 0,
    created_at      TEXT    DEFAULT (datetime('now'))
)
"""

PROJECT_ASSIGNMENTS_DDL = """
CREATE TABLE IF NOT EXISTS project_assignments (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    project_name    TEXT    NOT NULL,
    project_exec    TEXT,
    pm              TEXT,
    assistant_1     TEXT,
    assistant_2     TEXT,
    created_at      TEXT    DEFAULT (datetime('now')),
    UNIQUE(project_name)
)
"""

INVENTORY_DDL = """
CREATE TABLE IF NOT EXISTS inventory (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    item_name       TEXT    NOT NULL,
    quantity         TEXT,
    manufacturer    TEXT,
    location        TEXT,
    section         TEXT,
    updated_at      TEXT    DEFAULT (datetime('now')),
    created_at      TEXT    DEFAULT (datetime('now')),
    UNIQUE(item_name, location, section)
)
"""


# ---------------------------------------------------------------------------
# 1. PERDIEM-SHIRTS-EMAILS — Employee enrichment
# ---------------------------------------------------------------------------
def import_perdiem(db):
    """Update employees with email and phone from the PERDIEM sheet."""
    print("\n" + "=" * 60)
    print("1. PERDIEM-SHIRTS-EMAILS — Employee enrichment")
    print("=" * 60)

    wb = openpyxl.load_workbook(MISC_ITEMS_PATH, data_only=True)
    ws = wb["PERDIEM-SHIRTS-EMAILS"]
    employees = load_employees(db)

    # Check if email column exists (VPS has it, local may not)
    cols_info = db.execute("PRAGMA table_info(employees)").fetchall()
    has_email_col = any(c[1] == "email" for c in cols_info)

    updated_email = 0
    updated_phone = 0
    skipped = 0
    not_matched = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5, values_only=False):
        cell_vals = safe_cell_dict(row)

        name = clean(cell_vals.get("E"))
        email = clean(cell_vals.get("C"))
        phone_raw = cell_vals.get("D")

        # Skip header rows and empty rows
        if not name or name in ("#", "NAME", "SHIRT SIZE"):
            continue

        phone = normalize_phone_value(phone_raw)
        name_clean = clean_name(name)
        if not name_clean:
            continue

        emp, score = fuzzy_match_employee(name_clean, employees)
        if not emp:
            not_matched.append(f"  {name_clean} (best score: {score})")
            skipped += 1
            continue

        # Only fill in blanks — don't overwrite existing data
        updates = []
        params = []

        if email and not emp["email"] and has_email_col:
            updates.append("email = ?")
            params.append(email.lower())
            updated_email += 1

        if phone and not emp["phone_number"]:
            updates.append("phone_number = ?")
            params.append(phone)
            updated_phone += 1

        if updates:
            updates.append("updated_at = datetime('now')")
            params.append(emp["id"])
            sql = f"UPDATE employees SET {', '.join(updates)} WHERE id = ?"
            db.execute(sql, params)

    db.commit()
    wb.close()

    print(f"  Emails updated:    {updated_email}")
    print(f"  Phones updated:    {updated_phone}")
    print(f"  Not matched:       {skipped}")
    if not_matched:
        print("  Unmatched names:")
        for nm in not_matched:
            print(f"    {nm}")


# ---------------------------------------------------------------------------
# 2. Material Suppliers — Vendor data
# ---------------------------------------------------------------------------
def import_vendors(db):
    """Import material suppliers into vendors table."""
    print("\n" + "=" * 60)
    print("2. Material Suppliers — Vendor data")
    print("=" * 60)

    ensure_table(db, VENDORS_DDL)

    wb = openpyxl.load_workbook(MISC_ITEMS_PATH, data_only=True)
    ws = wb["Material Suppliers"]

    inserted = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6, values_only=False):
        cell_vals = safe_cell_dict(row)

        name = clean(cell_vals.get("A"))
        if not name or name == "COMPANY NAME":
            continue

        address = clean(cell_vals.get("B"))
        phone = clean(cell_vals.get("C"))
        contact = clean(cell_vals.get("D"))
        items = clean(cell_vals.get("E"))
        account_raw = clean(cell_vals.get("F"))

        # Normalize contact: n/a -> None
        if contact and contact.lower() == "n/a":
            contact = None

        has_account = 1 if account_raw and account_raw.upper() == "YES" else 0

        # Skip duplicates by name
        existing = db.execute(
            "SELECT id FROM vendors WHERE name = ?", (name,)
        ).fetchone()
        if existing:
            skipped += 1
            continue

        db.execute(
            """INSERT INTO vendors (name, address, phone, contact_name, items, has_account)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (name, address, phone, contact, items, has_account),
        )
        inserted += 1

    db.commit()
    wb.close()

    print(f"  Vendors inserted:  {inserted}")
    print(f"  Duplicates skipped:{skipped}")


# ---------------------------------------------------------------------------
# 3. PROJECT ASSIGNMENTS — Project assignment data
# ---------------------------------------------------------------------------
def import_project_assignments(db):
    """Import project-to-personnel assignments."""
    print("\n" + "=" * 60)
    print("3. PROJECT ASSIGNMENTS — Project assignment data")
    print("=" * 60)

    ensure_table(db, PROJECT_ASSIGNMENTS_DDL)

    wb = openpyxl.load_workbook(MISC_ITEMS_PATH, data_only=True)
    ws = wb["PROJECT ASSIGNMENTS"]

    inserted = 0
    skipped = 0

    # Known header/section keywords to skip
    skip_keywords = {
        "PROJECT", "ORLANDO & EAST COAST PROJECTS", "TAMPA / LAKELAND / WEST COAST",
        "TAMPA / LAKELAND / WEST COAST - not started yet", "NOT ASSIGNED",
    }
    # Known PM names that appear alone in column A as section headers
    pm_names = {"JUSTIN", "JAKE", "ROB", "RICHARD", "ZACK", "DOUG"}

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=5, values_only=False):
        cell_vals = safe_cell_dict(row)

        project_name = clean(cell_vals.get("A"))
        project_exec = clean(cell_vals.get("B"))
        pm = clean(cell_vals.get("C"))
        assistant_1 = clean(cell_vals.get("D"))
        assistant_2 = clean(cell_vals.get("E"))

        if not project_name:
            continue

        # Skip section headers and standalone PM names
        if project_name in skip_keywords or project_name.upper() in skip_keywords:
            continue
        if project_name.upper() in pm_names and not project_exec:
            continue

        # Must have at least a project exec or PM to be a real assignment
        if not project_exec and not pm:
            continue

        # Skip duplicate
        existing = db.execute(
            "SELECT id FROM project_assignments WHERE project_name = ?",
            (project_name,),
        ).fetchone()
        if existing:
            skipped += 1
            continue

        db.execute(
            """INSERT INTO project_assignments (project_name, project_exec, pm, assistant_1, assistant_2)
               VALUES (?, ?, ?, ?, ?)""",
            (project_name, project_exec, pm, assistant_1, assistant_2),
        )
        inserted += 1

    db.commit()
    wb.close()

    print(f"  Assignments inserted: {inserted}")
    print(f"  Duplicates skipped:   {skipped}")


# ---------------------------------------------------------------------------
# 4. SAFETY TRAINING — Certification data
# ---------------------------------------------------------------------------
def import_safety_training(db):
    """Import safety training X marks as certifications."""
    print("\n" + "=" * 60)
    print("4. SAFETY TRAINING — Certification updates")
    print("=" * 60)

    # Check if certification_types table exists (may be absent on local DB)
    table_check = db.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='certification_types'"
    ).fetchone()
    if not table_check:
        print("  SKIP — certification_types table not found (run on VPS)")
        return

    wb = openpyxl.load_workbook(MISC_ITEMS_PATH, data_only=True)
    ws = wb["SAFETY TRAINING"]
    employees = load_employees(db)

    # Load cert type slugs -> id mapping
    cert_types = {}
    for row in db.execute("SELECT id, slug, name FROM certification_types").fetchall():
        cert_types[row["slug"]] = row["id"]

    certs_created = 0
    certs_existed = 0
    not_matched = []

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=10, values_only=False):
        cell_vals = safe_cell_dict(row)

        name = clean(cell_vals.get("B"))
        if not name:
            continue

        # Skip legend rows
        if name in ("Training completed", "No training needed"):
            continue

        name_clean = clean_name(name)
        if not name_clean:
            continue

        emp, score = fuzzy_match_employee(name_clean, employees)
        if not emp:
            not_matched.append(f"  {name_clean} (best score: {score})")
            continue

        # Check each training column
        for col_letter, training_name in SAFETY_COLUMNS.items():
            cell_val = clean(cell_vals.get(col_letter))

            # Only 'X' means training completed
            if cell_val and cell_val.upper() == "X":
                slug = SAFETY_CERT_MAP.get(training_name)
                if not slug:
                    continue

                cert_type_id = cert_types.get(slug)
                if not cert_type_id:
                    print(f"  WARNING: cert type slug '{slug}' not found in DB")
                    continue

                # Check if cert already exists for this employee + type
                existing = db.execute(
                    """SELECT id FROM certifications
                       WHERE employee_id = ? AND cert_type_id = ? AND is_active = 1""",
                    (emp["id"], cert_type_id),
                ).fetchone()

                if existing:
                    certs_existed += 1
                    continue

                # Create the certification (no dates — just marking as held)
                db.execute(
                    """INSERT INTO certifications (employee_id, cert_type_id, notes, is_active)
                       VALUES (?, ?, ?, 1)""",
                    (emp["id"], cert_type_id, "Imported from SAFETY TRAINING spreadsheet"),
                )
                certs_created += 1

    db.commit()
    wb.close()

    print(f"  Certs created:     {certs_created}")
    print(f"  Already existed:   {certs_existed}")
    print(f"  Not matched:       {len(not_matched)}")
    if not_matched:
        print("  Unmatched names:")
        for nm in not_matched:
            print(f"    {nm}")


# ---------------------------------------------------------------------------
# 5. SHOP INVENTORY (Orlando) — from 2026 SCHEDULE.xlsx
# ---------------------------------------------------------------------------
def import_orlando_inventory(db):
    """Import Orlando shop inventory items."""
    print("\n" + "=" * 60)
    print("5. SHOP INVENTORY — Orlando Shop")
    print("=" * 60)

    ensure_table(db, INVENTORY_DDL)

    wb = openpyxl.load_workbook(SCHEDULE_PATH, data_only=True)
    # Sheet name has a trailing space
    ws = wb["SHOP INVENTORY "]

    inserted = 0
    skipped = 0
    current_section = None
    location = "Orlando Shop"

    # Skip rows: title, last updated, inventoried by, blank, header
    skip_values = {
        "ORLANDO SHOP INVENTORY", "LAST UPDATED:", "INVENTORIED BY:",
        "ITEM", "MATERIALS FOR JOBS",
    }

    # Section headers (no QTY, act as category groupings)
    section_headers = {
        "METAL", "FASTENERS", "BUCKETS", "BOOTS", "TPO", "MISC",
        "MATERIALS FOR JOBS",
    }

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3, values_only=False):
        cell_vals = safe_cell_dict(row)

        item_name = clean(cell_vals.get("A"))
        qty_raw = cell_vals.get("B")
        manufacturer = clean(cell_vals.get("C"))

        if not item_name:
            continue

        # Skip metadata rows
        if item_name in skip_values or item_name.startswith("LAST UPDATED") or item_name.startswith("INVENTORIED"):
            continue

        # Detect section headers
        if item_name.upper() in section_headers:
            current_section = item_name
            continue

        # Normalize quantity to string
        qty = clean(str(qty_raw)) if qty_raw is not None else None
        if qty == "0.0" or qty == "0":
            qty = "0"

        # Skip items from the "MATERIALS FOR JOBS" section (they're job-specific, not inventory)
        # These appear after row 58 with a JOB column in C
        # Detect: if item is in that section and C column looks like a job name
        # We handle this by checking if current_section is MATERIALS FOR JOBS
        if current_section == "MATERIALS FOR JOBS":
            continue

        # Check for duplicate
        existing = db.execute(
            "SELECT id FROM inventory WHERE item_name = ? AND location = ? AND section = ?",
            (item_name, location, current_section),
        ).fetchone()
        if existing:
            skipped += 1
            continue

        db.execute(
            """INSERT INTO inventory (item_name, quantity, manufacturer, location, section)
               VALUES (?, ?, ?, ?, ?)""",
            (item_name, qty, manufacturer, location, current_section),
        )
        inserted += 1

    db.commit()
    wb.close()

    print(f"  Items inserted:    {inserted}")
    print(f"  Duplicates skipped:{skipped}")


# ---------------------------------------------------------------------------
# 6. TAMPA SHOP INVENTORY — from MISC ITEMS.xlsx
# ---------------------------------------------------------------------------
def _parse_tampa_column_pair(ws, item_col, qty_col, section_label, db, location, manufacturer_col=None):
    """Parse a pair of (item, qty) columns from the Tampa inventory sheet.

    The Tampa sheet is a complex multi-column layout with many item/qty pairs
    spread across columns A-V. This helper handles one pair at a time.

    Returns (inserted, skipped) counts.
    """
    inserted = 0
    skipped = 0

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
        cell_map = safe_cell_dict(row)

        item_name = clean(cell_map.get(item_col))
        qty_raw = cell_map.get(qty_col)
        manufacturer = clean(cell_map.get(manufacturer_col)) if manufacturer_col else None

        if not item_name:
            continue

        # Skip sub-headers and metadata
        if item_name in ("QTY", "Buckets", "QTY "):
            continue
        # Skip rows that are just category headers used in context
        if item_name.startswith("Carlisle Insulfast") or item_name.startswith("Carlisle HP-X"):
            continue
        if item_name in ("Plates", "TPO rolls"):
            continue
        # Skip metadata
        if item_name.startswith("INVENTORY DATE") or item_name.startswith("Taken By"):
            continue
        # Skip insulation thickness markers (single letters/numbers used as codes)
        if item_name in ("Y", "X", "Q") and section_label == "Insulation":
            continue

        # Normalize quantity
        qty = None
        if qty_raw is not None:
            qty_str = str(qty_raw).strip()
            if qty_str and qty_str not in (" ",):
                qty = qty_str.replace(".0", "") if qty_str.endswith(".0") else qty_str
                if qty == "0":
                    qty = "0"

        # Dedup
        existing = db.execute(
            "SELECT id FROM inventory WHERE item_name = ? AND location = ? AND section = ?",
            (item_name, location, section_label),
        ).fetchone()
        if existing:
            skipped += 1
            continue

        db.execute(
            """INSERT INTO inventory (item_name, quantity, manufacturer, location, section)
               VALUES (?, ?, ?, ?, ?)""",
            (item_name, qty, manufacturer, location, section_label),
        )
        inserted += 1

    return inserted, skipped


def import_tampa_inventory(db):
    """Import Tampa shop inventory items.

    The Tampa sheet has a complex multi-section layout spread across many columns:
    - Cols A/B: Carlisle Insulfast Screws (ISO), Plates, TPO rolls, misc
    - Cols C/D: Carlisle HP-X Screws (TPO), Other TPO materials
    - Cols F/G: Hallway/Office items, DeWalt tools, storage
    - Cols H/I: Hand tools, power tools, blades
    - Cols J/K: Safety & rigging equipment
    - Cols M/N: Main wall shelving items
    - Cols O/P: Metal rack items (membrane, plastic, shrink wrap)
    - Cols S:   Bits and nut runners (no qty column)
    - Cols T/U: Metal rack continued (boots, stucco stop)
    """
    print("\n" + "=" * 60)
    print("6. TAMPA SHOP INVENTORY — Tampa Shop")
    print("=" * 60)

    ensure_table(db, INVENTORY_DDL)

    wb = openpyxl.load_workbook(MISC_ITEMS_PATH, data_only=True)
    ws = wb["TAMPA SHOP INVENTORY"]
    location = "Tampa Shop"

    total_inserted = 0
    total_skipped = 0

    # --- Column A/B: Screws (ISO), Plates, TPO rolls, misc ---
    section = "Screws & Plates"
    for row in ws.iter_rows(min_row=4, max_row=42, min_col=1, max_col=2, values_only=False):
        cell_map = safe_cell_dict(row)
        item = clean(cell_map.get("A"))
        qty_raw = cell_map.get("B")
        if not item:
            continue
        # Skip sub-headers
        if item in ("Plates", "TPO rolls", "INSULATION", "QTY"):
            continue
        if item.startswith("Carlisle Insulfast") or item.startswith("Taken By"):
            continue
        if item.startswith("INVENTORY DATE"):
            continue
        # Single letters Y, X, Q are insulation thickness markers
        if item in ("Y", "X", "Q"):
            continue
        # Numeric-only values (1.5, 2.0, 2.5) are insulation sizes, capture them
        if isinstance(cell_map.get("A"), (int, float)):
            raw_val = cell_map.get("A")
            # Insulation thickness entries
            item = f"Insulation {raw_val}\""
            section = "Insulation"

        qty = None
        if qty_raw is not None:
            qty = str(qty_raw).strip()
            if qty in (" ", "Buckets", "QTY"):
                qty = None
            elif qty.endswith(".0"):
                qty = qty.replace(".0", "")

        existing = db.execute(
            "SELECT id FROM inventory WHERE item_name = ? AND location = ? AND section = ?",
            (item, location, section),
        ).fetchone()
        if existing:
            total_skipped += 1
            continue

        db.execute(
            "INSERT INTO inventory (item_name, quantity, manufacturer, location, section) VALUES (?, ?, ?, ?, ?)",
            (item, qty, None, location, section),
        )
        total_inserted += 1

    # --- Column C/D: HP-X Screws (TPO) and Other TPO Materials ---
    section = "TPO Screws & Materials"
    for row in ws.iter_rows(min_row=4, max_row=43, min_col=3, max_col=4, values_only=False):
        cell_map = safe_cell_dict(row)
        item = clean(cell_map.get("C"))
        qty_raw = cell_map.get("D")
        if not item:
            continue
        if item.startswith("Carlisle HP-X") or item in ("OTHER TPO MTRLS", "QTY"):
            continue

        qty = None
        if qty_raw is not None:
            qty = str(qty_raw).strip()
            if qty in (" ", "QTY"):
                qty = None
            elif qty.endswith(".0"):
                qty = qty.replace(".0", "")

        existing = db.execute(
            "SELECT id FROM inventory WHERE item_name = ? AND location = ? AND section = ?",
            (item, location, section),
        ).fetchone()
        if existing:
            total_skipped += 1
            continue

        db.execute(
            "INSERT INTO inventory (item_name, quantity, manufacturer, location, section) VALUES (?, ?, ?, ?, ?)",
            (item, qty, None, location, section),
        )
        total_inserted += 1

    # --- Column F/G: Hallway/Office, DeWalt, Upper/Lower Storage ---
    section = "Hallway/Office & Storage"
    for row in ws.iter_rows(min_row=3, max_row=55, min_col=6, max_col=7, values_only=False):
        cell_map = safe_cell_dict(row)
        item = clean(cell_map.get("F"))
        qty_raw = cell_map.get("G")
        if not item:
            continue
        # Skip section labels (but we track them for context)
        if item in ("HALLWAY/OFFICE", "OFFICE", "DeWalt ", "DeWalt", "UPPER STORAGE", "LOWER STORAGE"):
            continue

        qty = None
        if qty_raw is not None:
            qty = str(qty_raw).strip()
            if qty in (" ",):
                qty = None
            elif qty.endswith(".0"):
                qty = qty.replace(".0", "")

        existing = db.execute(
            "SELECT id FROM inventory WHERE item_name = ? AND location = ? AND section = ?",
            (item, location, section),
        ).fetchone()
        if existing:
            total_skipped += 1
            continue

        db.execute(
            "INSERT INTO inventory (item_name, quantity, manufacturer, location, section) VALUES (?, ?, ?, ?, ?)",
            (item, qty, None, location, section),
        )
        total_inserted += 1

    # --- Column H/I: Tools ---
    section = "Tools"
    for row in ws.iter_rows(min_row=3, max_row=50, min_col=8, max_col=9, values_only=False):
        cell_map = safe_cell_dict(row)
        item = clean(cell_map.get("H"))
        qty_raw = cell_map.get("I")
        if not item:
            continue
        if item in ("Other tools",):
            continue

        qty = None
        if qty_raw is not None:
            qty = str(qty_raw).strip()
            if qty in (" ",):
                qty = None
            elif qty.endswith(".0"):
                qty = qty.replace(".0", "")

        existing = db.execute(
            "SELECT id FROM inventory WHERE item_name = ? AND location = ? AND section = ?",
            (item, location, section),
        ).fetchone()
        if existing:
            total_skipped += 1
            continue

        db.execute(
            "INSERT INTO inventory (item_name, quantity, manufacturer, location, section) VALUES (?, ?, ?, ?, ?)",
            (item, qty, None, location, section),
        )
        total_inserted += 1

    # --- Column J/K: Safety & Rigging ---
    section = "Safety & Rigging"
    for row in ws.iter_rows(min_row=3, max_row=30, min_col=10, max_col=11, values_only=False):
        cell_map = safe_cell_dict(row)
        item = clean(cell_map.get("J"))
        qty_raw = cell_map.get("K")
        if not item:
            continue
        if item in ("Safety & Rigging",):
            continue

        qty = None
        if qty_raw is not None:
            qty = str(qty_raw).strip()
            if qty in (" ",):
                qty = None
            elif qty.endswith(".0"):
                qty = qty.replace(".0", "")

        existing = db.execute(
            "SELECT id FROM inventory WHERE item_name = ? AND location = ? AND section = ?",
            (item, location, section),
        ).fetchone()
        if existing:
            total_skipped += 1
            continue

        db.execute(
            "INSERT INTO inventory (item_name, quantity, manufacturer, location, section) VALUES (?, ?, ?, ?, ?)",
            (item, qty, None, location, section),
        )
        total_inserted += 1

    # --- Column M/N: Main Wall Shelving ---
    section = "Main Wall Shelving"
    for row in ws.iter_rows(min_row=3, max_row=45, min_col=13, max_col=14, values_only=False):
        cell_map = safe_cell_dict(row)
        item = clean(cell_map.get("M"))
        qty_raw = cell_map.get("N")
        if not item:
            continue
        # Skip section labels and to-do items
        skip_items = {
            "METAL RACK (right to left from garage door)", "METAL RACK",
            "MAIN WALL SHELVING", "ROLLAR CAGES - ROLLER NAPS - RAGS",
            "Shop to do", "Clean small shop", "Pallet jack on big shop side",
            "TPO rolls to big shop", "Spray rig to big shop", "Simplex to big shop",
            "TPO metal to big shop", "Broken generators tossed out", "Big shop",
            "Rags, roller cages, naps move", "Water pallets moved to Rag location",
            "Bathroom painted",
        }
        # Also skip "NEED TO ORDER" and supply list items in N column
        if item in skip_items:
            continue

        qty = None
        if qty_raw is not None:
            qty_str = str(qty_raw).strip()
            # N column sometimes has "NEED TO ORDER", "SCISSORS" etc. — not quantities
            if qty_str and qty_str not in (" ",) and not qty_str.isalpha():
                qty = qty_str.replace(".0", "") if qty_str.endswith(".0") else qty_str

        existing = db.execute(
            "SELECT id FROM inventory WHERE item_name = ? AND location = ? AND section = ?",
            (item, location, section),
        ).fetchone()
        if existing:
            total_skipped += 1
            continue

        db.execute(
            "INSERT INTO inventory (item_name, quantity, manufacturer, location, section) VALUES (?, ?, ?, ?, ?)",
            (item, qty, None, location, section),
        )
        total_inserted += 1

    # --- Column O/P: Metal Rack items (caulk, membrane, etc.) ---
    section = "Metal Rack"
    for row in ws.iter_rows(min_row=3, max_row=30, min_col=15, max_col=16, values_only=False):
        cell_map = safe_cell_dict(row)
        item = clean(cell_map.get("O"))
        qty_raw = cell_map.get("P")
        if not item:
            continue
        if item in ("MEMBRANE",):
            continue

        qty = None
        if qty_raw is not None:
            qty = str(qty_raw).strip()
            if qty in (" ",):
                qty = None
            elif qty.endswith(".0"):
                qty = qty.replace(".0", "")

        existing = db.execute(
            "SELECT id FROM inventory WHERE item_name = ? AND location = ? AND section = ?",
            (item, location, section),
        ).fetchone()
        if existing:
            total_skipped += 1
            continue

        db.execute(
            "INSERT INTO inventory (item_name, quantity, manufacturer, location, section) VALUES (?, ?, ?, ?, ?)",
            (item, qty, None, location, section),
        )
        total_inserted += 1

    # --- Column S: Bits & Nut Runners (no qty column) ---
    section = "Bits & Fastener Tools"
    for row in ws.iter_rows(min_row=25, max_row=35, min_col=19, max_col=19, values_only=False):
        cell_map = safe_cell_dict(row)
        item = clean(cell_map.get("S"))
        if not item:
            continue

        existing = db.execute(
            "SELECT id FROM inventory WHERE item_name = ? AND location = ? AND section = ?",
            (item, location, section),
        ).fetchone()
        if existing:
            total_skipped += 1
            continue

        db.execute(
            "INSERT INTO inventory (item_name, quantity, manufacturer, location, section) VALUES (?, ?, ?, ?, ?)",
            (item, None, None, location, section),
        )
        total_inserted += 1

    # --- Column T/U: Metal rack continued (boots, stucco stop, AC Metal) ---
    section = "Metal Rack - Boots & Flashing"
    # Row 3 has headers in T/U, row 4 has quantities
    # Then row 7 has more headers, row 8 has quantities
    # Then rows 12-16 have individual items with quantities in U
    metal_rack_items = [
        # (item_name_col, qty_col, row_range_start, row_range_end)
    ]
    for row in ws.iter_rows(min_row=3, max_row=20, min_col=20, max_col=22, values_only=False):
        cell_map = safe_cell_dict(row)
        item_t = clean(cell_map.get("T"))
        item_u_name = clean(cell_map.get("U"))
        qty_u = cell_map.get("U")

        # Handle T column items with their quantities (often in the next row)
        if item_t and not str(cell_map.get("T", "")).replace(".", "").isdigit():
            # T has an item name, check if U has a quantity
            if isinstance(qty_u, (int, float)):
                qty = str(int(qty_u)) if isinstance(qty_u, float) and qty_u == int(qty_u) else str(qty_u)
            else:
                qty = None

            existing = db.execute(
                "SELECT id FROM inventory WHERE item_name = ? AND location = ? AND section = ?",
                (item_t, location, section),
            ).fetchone()
            if not existing:
                db.execute(
                    "INSERT INTO inventory (item_name, quantity, manufacturer, location, section) VALUES (?, ?, ?, ?, ?)",
                    (item_t, qty, None, location, section),
                )
                total_inserted += 1
            else:
                total_skipped += 1

    # Specific named items from T/U that have quantities in U column (rows 12-16)
    boot_items = {
        12: ("TPO BOOTS - SPLIT PIPE", "U"),
        14: ("TPO universal Boot", "U"),
        16: ("TPO one way", "U"),
    }
    for row_num, (item_name, qty_col) in boot_items.items():
        for row in ws.iter_rows(min_row=row_num, max_row=row_num, values_only=False):
            cell_map = safe_cell_dict(row)
            qty_raw = cell_map.get(qty_col)
            qty = None
            if qty_raw is not None and isinstance(qty_raw, (int, float)):
                qty = str(int(qty_raw))

            existing = db.execute(
                "SELECT id FROM inventory WHERE item_name = ? AND location = ? AND section = ?",
                (item_name, location, section),
            ).fetchone()
            if not existing:
                db.execute(
                    "INSERT INTO inventory (item_name, quantity, manufacturer, location, section) VALUES (?, ?, ?, ?, ?)",
                    (item_name, qty, None, location, section),
                )
                total_inserted += 1
            else:
                total_skipped += 1

    # --- Downspouts from J/K columns (rows 34-36) ---
    section = "Downspouts"
    for row in ws.iter_rows(min_row=34, max_row=36, min_col=10, max_col=11, values_only=False):
        cell_map = safe_cell_dict(row)
        item = clean(cell_map.get("J"))
        qty_raw = cell_map.get("K")
        if not item:
            continue

        qty = None
        if qty_raw is not None:
            qty = str(qty_raw).strip()
            if qty.endswith(".0"):
                qty = qty.replace(".0", "")

        existing = db.execute(
            "SELECT id FROM inventory WHERE item_name = ? AND location = ? AND section = ?",
            (item, location, section),
        ).fetchone()
        if not existing:
            db.execute(
                "INSERT INTO inventory (item_name, quantity, manufacturer, location, section) VALUES (?, ?, ?, ?, ?)",
                (item, qty, None, location, section),
            )
            total_inserted += 1
        else:
            total_skipped += 1

    db.commit()
    wb.close()

    print(f"  Items inserted:    {total_inserted}")
    print(f"  Duplicates skipped:{total_skipped}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    global MISC_ITEMS_PATH, SCHEDULE_PATH

    import argparse
    parser = argparse.ArgumentParser(description="Import spreadsheet data into CrewLedger")
    parser.add_argument("--misc", default=None, help="Path to MISC ITEMS.xlsx")
    parser.add_argument("--schedule", default=None, help="Path to 2026 SCHEDULE.xlsx")
    parser.add_argument("--db", default=None, help="Path to database file")
    parser.add_argument("--dry-run", action="store_true", help="Parse and print without writing")
    args = parser.parse_args()

    # Override module-level paths if provided
    if args.misc:
        MISC_ITEMS_PATH = args.misc
    if args.schedule:
        SCHEDULE_PATH = args.schedule

    print("=" * 60)
    print("CrewLedger Spreadsheet Data Import")
    print("=" * 60)

    # Verify files exist
    for path, label in [(MISC_ITEMS_PATH, "MISC ITEMS.xlsx"), (SCHEDULE_PATH, "2026 SCHEDULE.xlsx")]:
        if not os.path.exists(path):
            print(f"ERROR: {label} not found at {path}")
            sys.exit(1)
        print(f"  Found: {path}")

    db = get_db(args.db)

    try:
        import_perdiem(db)
        import_vendors(db)
        import_project_assignments(db)
        import_safety_training(db)
        import_orlando_inventory(db)
        import_tampa_inventory(db)

        print("\n" + "=" * 60)
        print("IMPORT COMPLETE")
        print("=" * 60)

        # Print summary counts (guard against missing tables)
        def safe_count(table):
            try:
                return db.execute(f"SELECT COUNT(*) as c FROM {table}").fetchone()["c"]
            except Exception:
                return "N/A"

        print(f"\n  Total vendors:            {safe_count('vendors')}")
        print(f"  Total project assignments:{safe_count('project_assignments')}")
        print(f"  Total inventory items:    {safe_count('inventory')}")
        print(f"  Total certifications:     {safe_count('certifications')}")
        print()

    finally:
        db.close()


if __name__ == "__main__":
    main()
